"""
Excel I/O utilities for reading and writing engagement workbooks.

ExcelWorkbookManager wraps openpyxl and provides the narrow interface the
rest of the framework needs: key-value config reads, dict-based row reads,
and dict-based row writes.  Direct openpyxl access is available via
get_sheet() for cases that need lower-level control (e.g., reading headers
to infer an output schema).

Expected workbook structure
---------------------------
CONFIG sheet  — column A = key name, column B = value. Used for all
                engagement metadata (EngagementID, dates, approvers, etc.).
TAG_ENGINE    — tabular; first row = headers, subsequent rows = tag defs.
QA            — tabular; first row = headers, subsequent rows = QA rules.
OUTPUT        — tabular; first row = output column headers.
AUDIT_LOG     — written by AuditLog; created on demand if missing.

Thread safety
-------------
openpyxl workbooks are NOT thread-safe.  Create one ExcelWorkbookManager
per thread/process.  The class supports use as a context manager to ensure
the workbook is always closed after use:

    with ExcelWorkbookManager("engagement.xlsx") as mgr:
        value = mgr.read_config_value("EngagementID")
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from audit_framework.exceptions import WorksheetMissingError


class ExcelWorkbookManager:
    """Manages reading and writing of Excel engagement workbooks via openpyxl."""

    def __init__(self, file_path: str | Path) -> None:
        """
        Load the workbook at file_path into memory.

        Args:
            file_path: Absolute or relative path to the .xlsx / .xlsm file.

        Raises:
            FileNotFoundError: If the file does not exist at the given path.
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.file_path}")

        # data_only=True reads cached formula results rather than formula strings,
        # which is what we want when reading extracted data cells.
        self.workbook = load_workbook(self.file_path)

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """
        Return a worksheet by name, raising a typed error if absent.

        Args:
            sheet_name: Case-sensitive sheet name.

        Returns:
            The openpyxl Worksheet object.

        Raises:
            WorksheetMissingError: If the sheet doesn't exist in this workbook.
        """
        if sheet_name not in self.workbook.sheetnames:
            raise WorksheetMissingError(
                f"Sheet '{sheet_name}' not found in workbook. "
                f"Available sheets: {self.workbook.sheetnames}"
            )
        return self.workbook[sheet_name]

    def sheet_exists(self, sheet_name: str) -> bool:
        """Return True if the named sheet exists in the workbook."""
        return sheet_name in self.workbook.sheetnames

    def read_cell(self, sheet_name: str, row: int, col: int) -> Any:
        """
        Read a single cell by 1-based row and column indices.

        Returns:
            The cell's value, or None if the cell is empty.
        """
        ws = self.get_sheet(sheet_name)
        return ws.cell(row, col).value

    def read_cell_by_ref(self, sheet_name: str, cell_ref: str) -> Any:
        """
        Read a single cell by Excel reference string (e.g. 'B4').

        Returns:
            The cell's value, or None if the cell is empty.
        """
        ws = self.get_sheet(sheet_name)
        return ws[cell_ref].value

    def write_cell(self, sheet_name: str, row: int, col: int, value: Any) -> None:
        """Write a value to a cell identified by 1-based row and column indices."""
        ws = self.get_sheet(sheet_name)
        ws.cell(row, col, value)

    def write_cell_by_ref(self, sheet_name: str, cell_ref: str, value: Any) -> None:
        """Write a value to a cell identified by Excel reference string (e.g. 'B4')."""
        ws = self.get_sheet(sheet_name)
        ws[cell_ref] = value

    def read_config_value(self, key: str) -> Optional[str]:
        """
        Look up a value from the CONFIG sheet by its key name.

        The CONFIG sheet uses a two-column layout:
            Column A — key name  (e.g. "EngagementID")
            Column B — value     (e.g. "ENG-2024-001")

        The search is case-insensitive so "engagementid" matches "EngagementID".

        Args:
            key: The key to look up.

        Returns:
            The value from column B as a raw openpyxl value, or None if the
            key is not present in the sheet.
        """
        ws = self.get_sheet("CONFIG")

        for row_idx in range(1, ws.max_row + 1):
            cell_key = ws.cell(row_idx, 1).value
            if cell_key and str(cell_key).strip().lower() == key.lower():
                return ws.cell(row_idx, 2).value

        return None

    def read_rows_as_dicts(
        self, sheet_name: str, header_row: int = 1
    ) -> list[dict[str, Any]]:
        """
        Read a tabular sheet into a list of dicts keyed by header values.

        Rows where every cell is None are skipped (handles trailing empty rows
        that openpyxl sometimes reports within max_row).  Columns without a
        header cell are given a synthetic name ("Column_N") so no data is lost.

        Args:
            sheet_name: Name of the sheet to read.
            header_row:  1-based row number containing column headers.

        Returns:
            List of dicts; each dict represents one data row.
        """
        ws = self.get_sheet(sheet_name)

        # Build the header list from the header row
        headers: list[str] = []
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col_idx).value
            headers.append(str(header).strip() if header else f"Column_{col_idx}")

        # Read data rows, skipping fully-empty ones
        rows: list[dict[str, Any]] = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data: dict[str, Any] = {}
            has_data = False
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value is not None:
                    has_data = True
                row_data[header] = cell_value

            if has_data:
                rows.append(row_data)

        return rows

    def write_rows_from_dicts(
        self,
        sheet_name: str,
        data: list[dict[str, Any]],
        headers: Optional[list[str]] = None,
        start_row: int = 1,
    ) -> None:
        """
        Write a list of dicts to a sheet, with headers in the first row.

        Args:
            sheet_name: Target sheet (must already exist).
            data:        Rows to write; each is a dict of {column: value}.
            headers:     Column order. If None, uses keys from the first row.
            start_row:   1-based row to write the header line.
        """
        ws = self.get_sheet(sheet_name)

        if not data:
            return

        if headers is None:
            headers = list(data[0].keys())

        # Write column headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(start_row, col_idx, header)

        # Write data rows immediately below the header
        for row_offset, row_data in enumerate(data, 1):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(start_row + row_offset, col_idx, row_data.get(header))

    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> Worksheet:
        """
        Create a new worksheet in the workbook.

        Args:
            sheet_name: Name for the new sheet (must be unique).
            index:       Optional position in the tab order (0 = leftmost).

        Returns:
            The newly created Worksheet.

        Raises:
            ValueError: If a sheet with that name already exists.
        """
        if sheet_name in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        return (
            self.workbook.create_sheet(sheet_name, index)
            if index is not None
            else self.workbook.create_sheet(sheet_name)
        )

    def save(self) -> None:
        """Persist all in-memory changes back to the original file path."""
        self.workbook.save(self.file_path)

    def close(self) -> None:
        """Close the workbook and release the file handle."""
        self.workbook.close()

    def __enter__(self) -> "ExcelWorkbookManager":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()
