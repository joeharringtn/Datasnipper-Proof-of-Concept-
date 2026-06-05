"""
Excel utilities for reading and writing engagement workbooks.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from audit_framework.exceptions import WorksheetMissingError


class ExcelWorkbookManager:
    """Manages reading and writing Excel engagement workbooks."""

    def __init__(self, file_path: str | Path) -> None:
        """
        Initialize workbook manager.

        Args:
            file_path: Path to Excel workbook

        Raises:
            FileNotFoundError: If workbook file doesn't exist
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.file_path}")

        self.workbook = load_workbook(self.file_path)

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """
        Get worksheet by name.

        Args:
            sheet_name: Name of worksheet

        Returns:
            Worksheet object

        Raises:
            WorksheetMissingError: If worksheet doesn't exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise WorksheetMissingError(
                f"Sheet '{sheet_name}' not found in workbook. "
                f"Available sheets: {self.workbook.sheetnames}"
            )
        return self.workbook[sheet_name]

    def sheet_exists(self, sheet_name: str) -> bool:
        """Check if worksheet exists."""
        return sheet_name in self.workbook.sheetnames

    def read_cell(
        self, sheet_name: str, row: int, col: int
    ) -> Any:
        """
        Read value from cell.

        Args:
            sheet_name: Name of worksheet
            row: 1-based row number
            col: 1-based column number

        Returns:
            Cell value (None if empty)
        """
        ws = self.get_sheet(sheet_name)
        return ws.cell(row, col).value

    def read_cell_by_ref(self, sheet_name: str, cell_ref: str) -> Any:
        """
        Read value from cell using cell reference (e.g., 'A1').

        Args:
            sheet_name: Name of worksheet
            cell_ref: Cell reference string

        Returns:
            Cell value (None if empty)
        """
        ws = self.get_sheet(sheet_name)
        return ws[cell_ref].value

    def write_cell(
        self, sheet_name: str, row: int, col: int, value: Any
    ) -> None:
        """
        Write value to cell.

        Args:
            sheet_name: Name of worksheet
            row: 1-based row number
            col: 1-based column number
            value: Value to write
        """
        ws = self.get_sheet(sheet_name)
        ws.cell(row, col, value)

    def write_cell_by_ref(self, sheet_name: str, cell_ref: str, value: Any) -> None:
        """
        Write value to cell using cell reference (e.g., 'A1').

        Args:
            sheet_name: Name of worksheet
            cell_ref: Cell reference string
            value: Value to write
        """
        ws = self.get_sheet(sheet_name)
        ws[cell_ref] = value

    def read_config_value(self, key: str) -> Optional[str]:
        """
        Read configuration value from CONFIG sheet.

        CONFIG sheet format: Column A = keys, Column B = values

        Args:
            key: Configuration key (case-insensitive)

        Returns:
            Configuration value or None if not found
        """
        ws = self.get_sheet("CONFIG")
        last_row = ws.max_row

        for row_idx in range(1, last_row + 1):
            cell_key = ws.cell(row_idx, 1).value
            if cell_key and str(cell_key).strip().lower() == key.lower():
                return ws.cell(row_idx, 2).value

        return None

    def read_rows_as_dicts(
        self, sheet_name: str, header_row: int = 1
    ) -> list[dict[str, Any]]:
        """
        Read worksheet as list of dictionaries.

        Args:
            sheet_name: Name of worksheet
            header_row: Row number containing headers (1-based, default 1)

        Returns:
            List of dictionaries where keys are headers
        """
        ws = self.get_sheet(sheet_name)

        # Read headers
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col_idx).value
            if header:
                headers.append(str(header).strip())
            else:
                headers.append(f"Column_{col_idx}")

        # Read data rows
        rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = {}
            has_data = False
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value is not None:
                    has_data = True
                row_data[header] = cell_value

            if has_data:
                rows.append(row_data)

        return rows

    def write_rows_from_dicts(
        self,
        sheet_name: str,
        data: list[dict[str, Any]],
        headers: Optional[list[str]] = None,
        start_row: int = 1,
    ) -> None:
        """
        Write list of dictionaries to worksheet.

        Args:
            sheet_name: Name of worksheet
            data: List of dictionaries
            headers: Column headers (if None, uses dict keys from first row)
            start_row: Starting row number (1-based)
        """
        ws = self.get_sheet(sheet_name)

        if not data:
            return

        # Determine headers
        if headers is None:
            headers = list(data[0].keys())

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(start_row, col_idx, header)

        # Write data rows
        for row_offset, row_data in enumerate(data, 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                ws.cell(start_row + row_offset, col_idx, value)

    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> Worksheet:
        """
        Create new worksheet.

        Args:
            sheet_name: Name for new worksheet
            index: Position in workbook (optional)

        Returns:
            New Worksheet object
        """
        if sheet_name in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        if index is not None:
            return self.workbook.create_sheet(sheet_name, index)
        else:
            return self.workbook.create_sheet(sheet_name)

    def save(self) -> None:
        """Save workbook."""
        self.workbook.save(self.file_path)

    def close(self) -> None:
        """Close workbook."""
        self.workbook.close()

    def __enter__(self) -> "ExcelWorkbookManager":
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """Context manager exit."""
        self.close()
