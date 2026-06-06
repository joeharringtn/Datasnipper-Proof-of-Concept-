"""Tests for ConfigManager module."""

import pytest

from audit_framework.config_manager import ConfigManager
from audit_framework.exceptions import ConfigError, WorksheetMissingError
from audit_framework.models import EngagementType


def test_config_manager_requires_config_sheet(tmp_path):
    """Test that ConfigManager requires CONFIG sheet."""
    from openpyxl import Workbook

    # Create workbook without CONFIG sheet
    wb = Workbook()
    wb.remove(wb.active)
    workbook_path = tmp_path / "test.xlsx"
    wb.save(workbook_path)

    # Should raise WorksheetMissingError
    with pytest.raises(WorksheetMissingError):
        ConfigManager(workbook_path)


def test_config_manager_missing_file():
    """Test that ConfigManager fails on missing file."""
    with pytest.raises(FileNotFoundError):
        ConfigManager("/nonexistent/path/workbook.xlsx")


def test_config_load_missing_required_field(tmp_path):
    """Test loading configuration with missing required field."""
    from openpyxl import Workbook

    # Create workbook with CONFIG sheet but missing EngagementID
    wb = Workbook()
    ws = wb.active
    ws.title = "CONFIG"
    ws["A1"] = "EngagementType"
    ws["B1"] = "Cash"
    workbook_path = tmp_path / "test.xlsx"
    wb.save(workbook_path)

    config_mgr = ConfigManager(workbook_path)
    with pytest.raises(ConfigError):
        config_mgr.load_config()


def test_config_validation():
    """Test configuration validation logic."""
    from audit_framework.models import ConfigObject, EngagementType
    from datetime import date

    config = ConfigObject(
        engagement_id="TEST-001",
        engagement_type=EngagementType.CASH,
        period_start_date=date(2026, 1, 1),
        period_end_date=date(2026, 12, 31),
        lead_auditor="John Doe",
        client_name="Test Company",
    )

    assert config.engagement_id == "TEST-001"
    assert config.engagement_type == EngagementType.CASH
