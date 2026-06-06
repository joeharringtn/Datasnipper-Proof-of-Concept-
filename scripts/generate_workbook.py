"""
Generate a DataSnipper-ready Excel engagement workbook.

When --source-pdf points to a PDF that exists on disk, the script reads
the PDF, shows you a numbered list of text found on the page, and you
pick which terms to search for.  The workbook is then generated with
DS_SEARCH tags built from your selections — no manual editing required.

If the PDF is not on the current machine (e.g. generating on a machine
without the file), placeholder tags are used instead.

DataSnipper tag format (v4.0+)
-------------------------------
  DS_SEARCH[filename|pageNumber|query]
  DS_COORDS[filename|pageNumber|x1|y1|x2|y2]

  The workbook filename MUST end in _(ds) — e.g. engagement_(ds).xlsx

Usage
-----
  # Interactive: reads your PDF, prompts for search term selection
  python scripts/generate_workbook.py --source-pdf "C:\\Audits\\invoice.pdf"

  # Non-interactive: pass terms directly, skip the prompt
  python scripts/generate_workbook.py \\
      --source-pdf "C:\\Audits\\invoice.pdf" \\
      --terms "Invoice Number,Total Amount,Due Date"

  # Specify which page to extract candidates from (default: 1)
  python scripts/generate_workbook.py --source-pdf "invoice.pdf" --page 2

  # Override output location
  python scripts/generate_workbook.py \\
      --source-pdf "C:\\Audits\\invoice.pdf" \\
      --output "C:\\Audits\\client_ap_(ds).xlsx"
"""

import argparse
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))

from audit_framework.tag_builder import TagBuilder
from audit_framework.models import (
    ExtractionMethod,
    TagDefinition,
    QARule,
    RuleType,
    FailAction,
    Severity,
)


# ── Styling ──────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
TAG_COL_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")


def _style_header_row(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_fit_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)


# ── PDF extraction ────────────────────────────────────────────────────────────

def extract_pdf_candidates(pdf_path: Path, page: int = 1) -> list[str]:
    """
    Extract short text lines from a PDF page as DS_SEARCH query candidates.

    Returns an empty list if pdfplumber is unavailable or the page is blank.
    """
    try:
        import pdfplumber
    except ImportError:
        return []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page < 1 or page > len(pdf.pages):
                print(f"  Warning: page {page} out of range (PDF has {len(pdf.pages)} pages)")
                return []
            text = pdf.pages[page - 1].extract_text() or ""
    except Exception as e:
        print(f"  Warning: could not read PDF ({e})")
        return []

    candidates = []
    for line in text.splitlines():
        line = line.strip()
        # Keep short lines — these are likely labels, not paragraph text
        if line and len(line) <= 80:
            candidates.append(line)
    return candidates


def pick_terms_interactively(candidates: list[str]) -> list[str]:
    """
    Show numbered candidates and return the user's selection.

    The user can enter comma-separated numbers ('1,3,5'), 'all', or type
    custom terms manually.
    """
    print("\nText found in PDF (candidates for DS_SEARCH queries):")
    print("-" * 60)
    for i, c in enumerate(candidates, 1):
        print(f"  {i:3}. {c}")
    print("-" * 60)
    print("Enter numbers to use (e.g. '1,3,5'), 'all' for everything,")
    print("or type custom terms separated by commas:")

    raw = input("> ").strip()

    if not raw:
        return []

    if raw.lower() == "all":
        return candidates

    # Try to parse as numbers first
    if re.match(r'^[\d,\s]+$', raw):
        try:
            indices = [int(x.strip()) - 1 for x in raw.split(",") if x.strip()]
            selected = [candidates[i] for i in indices if 0 <= i < len(candidates)]
            if selected:
                return selected
        except (ValueError, IndexError):
            pass

    # Otherwise treat as comma-separated custom terms
    return [t.strip() for t in raw.split(",") if t.strip()]


def make_dynamic_tags(source_doc: str, terms: list[str], pages: list[int]) -> list[TagDefinition]:
    """
    Build TagDefinitions from search terms × pages.

    One tag is created per (term, page) combination so DataSnipper searches
    every requested page.  Field names get a _p{n} suffix when multiple pages
    are requested so each tag has a unique name.
    """
    tags = []
    counter = 1
    multi_page = len(pages) > 1
    for term in terms:
        for page in pages:
            field_name = re.sub(r"[^a-z0-9]+", "_", term.lower()).strip("_") or f"field_{counter}"
            if multi_page:
                field_name = f"{field_name}_p{page}"
            tags.append(TagDefinition(
                tag_id=f"TAG_{counter:03d}_S",
                extraction_method=ExtractionMethod.DS_SEARCH,
                field_name=field_name,
                source_document=source_doc,
                search_page=page,
                search_keywords=term,
                notes=f"Search page {page} for '{term}'",
            ))
            counter += 1
    return tags


# ── Fallback sample data ──────────────────────────────────────────────────────

def make_sample_ap_tags(source_doc: str) -> list[TagDefinition]:
    """
    Placeholder AP tags used when the PDF is not available on this machine.
    Replace search_keywords values with text that actually appears in your PDF.
    """
    placeholders = [
        ("AP_InvoiceNum_S",   "invoice_number",       "Invoice Number"),
        ("AP_VendorName_S",   "vendor_name",          "Bill From"),
        ("AP_InvoiceDate_S",  "invoice_date",         "Invoice Date"),
        ("AP_InvoiceTotal_S", "invoice_total",        "Total Amount"),
        ("AP_PoNumber_S",     "po_number",            "PO Number"),
        ("AP_DueDate_S",      "due_date",             "Due Date"),
    ]
    return [
        TagDefinition(
            tag_id=tag_id,
            extraction_method=ExtractionMethod.DS_SEARCH,
            field_name=field_name,
            source_document=source_doc,
            search_page=1,
            search_keywords=keyword,
            notes="PLACEHOLDER - update search_keywords to match your PDF",
        )
        for tag_id, field_name, keyword in placeholders
    ]


def make_sample_qa_rules() -> list[QARule]:
    return [
        QARule(
            rule_id="QA_001",
            rule_name="Invoice total must be positive",
            field_name="invoice_total",
            rule_type=RuleType.RANGE,
            rule_definition="0.01:99999999",
            fail_action=FailAction.BLOCK,
            severity=Severity.CRITICAL,
            priority=1,
        ),
        QARule(
            rule_id="QA_002",
            rule_name="PO number format",
            field_name="po_number",
            rule_type=RuleType.FORMAT,
            rule_definition=r"^[A-Z0-9\-]{4,20}$",
            fail_action=FailAction.FLAG,
            severity=Severity.MEDIUM,
            priority=2,
        ),
    ]


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_config_sheet(wb: Workbook, config: dict) -> None:
    ws = wb.create_sheet("CONFIG")
    ws["A1"] = "Engagement Configuration"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    headers = ["Key", "Value", "DataType", "Scope", "Description", "Required"]
    for col, h in enumerate(headers, 1):
        ws.cell(3, col, h)
    _style_header_row(ws, 3, len(headers))

    rows = [
        ("EngagementID",     config.get("engagement_id", ""),       "String", "Global", "Unique engagement identifier",               "Yes"),
        ("EngagementType",   config.get("engagement_type", "AP"),   "String", "Global", "Audit area (Cash/AR/AP/Contracts/Inventory)", "Yes"),
        ("PeriodStartDate",  str(config.get("period_start", "")),   "Date",   "Global", "Start date for the engagement period",        "Yes"),
        ("PeriodEndDate",    str(config.get("period_end", "")),     "Date",   "Global", "End date for the engagement period",          "Yes"),
        ("LeadAuditor",      config.get("lead_auditor", ""),        "String", "Global", "Primary audit lead",                         "Yes"),
        ("ClientName",       config.get("client_name", ""),         "String", "Global", "Client name",                                "Yes"),
        ("ClientContact",    config.get("client_contact", ""),      "String", "Global", "Client contact name",                        "No"),
        ("FrameworkVersion", "1.0",                                  "String", "Global", "Version of this framework",                  "No"),
        ("WorkflowType",     config.get("workflow_type", "SIMPLE"), "String", "Global", "Approval workflow (SIMPLE/MULTI_LEVEL/NONE)", "No"),
    ]
    for row_idx, row_data in enumerate(rows, 4):
        for col, value in enumerate(row_data, 1):
            ws.cell(row_idx, col, value)
    _auto_fit_columns(ws)


def build_tag_engine_sheet(wb: Workbook, tag_definitions: list[TagDefinition]) -> None:
    """
    The SourceTag column (green header) holds DS_SEARCH[...] / DS_COORDS[...]
    strings.  DataSnipper scans the workbook on open and processes every cell
    it finds containing one of those strings.
    """
    ws = wb.create_sheet("TAG_ENGINE")

    headers = [
        "TagID", "FieldName", "ExtractionMethod",
        "SourceDocument", "SearchPage", "Query",
        "CoordPage", "X1", "Y1", "X2", "Y2",
        "Notes", "SourceTag",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header_row(ws, 1, len(headers))
    ws.cell(1, headers.index("SourceTag") + 1).fill = TAG_COL_FILL

    for row_idx, tag in enumerate(tag_definitions, 2):
        try:
            tag_string = TagBuilder.build_tag(tag)
        except Exception as e:
            tag_string = f"BUILD_ERROR: {e}"

        x2 = (tag.coord_x + tag.coord_width) if tag.coord_width else ""
        y2 = (tag.coord_y + tag.coord_height) if tag.coord_height else ""
        is_search = tag.extraction_method == ExtractionMethod.DS_SEARCH
        is_coords = tag.extraction_method == ExtractionMethod.DS_COORDS

        values = [
            tag.tag_id, tag.field_name, tag.extraction_method.value,
            tag.source_document,
            tag.search_page if is_search else "",
            (tag.search_keywords or "") if is_search else "",
            tag.coord_page if is_coords else "",
            tag.coord_x    if is_coords else "",
            tag.coord_y    if is_coords else "",
            x2             if is_coords else "",
            y2             if is_coords else "",
            tag.notes or "",
            tag_string,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row_idx, col, value)

    _auto_fit_columns(ws)


def build_qa_sheet(wb: Workbook, rules: list[QARule]) -> None:
    ws = wb.create_sheet("QA")
    headers = ["RuleID", "RuleName", "FieldName", "RuleType",
               "RuleDefinition", "FailAction", "Severity", "Priority"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header_row(ws, 1, len(headers))
    for row_idx, rule in enumerate(rules, 2):
        values = [rule.rule_id, rule.rule_name, rule.field_name,
                  rule.rule_type.value, rule.rule_definition,
                  rule.fail_action.value, rule.severity.value, rule.priority]
        for col, value in enumerate(values, 1):
            ws.cell(row_idx, col, value)
    _auto_fit_columns(ws)


def build_output_sheet(wb: Workbook, output_columns: list[str]) -> None:
    ws = wb.create_sheet("OUTPUT")
    for col, name in enumerate(output_columns, 1):
        ws.cell(1, col, name)
    _style_header_row(ws, 1, len(output_columns))
    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)


def build_audit_log_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("AUDIT_LOG")
    headers = ["Timestamp", "EventType", "Message", "User", "Module",
               "RowID", "FieldName", "OldValue", "NewValue", "Details"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header_row(ws, 1, len(headers))
    _auto_fit_columns(ws)


# ── Entry point ───────────────────────────────────────────────────────────────

def generate(output_path: Path, tags: list[TagDefinition], config: dict) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    rules = make_sample_qa_rules()
    output_columns = [t.field_name for t in tags]

    build_config_sheet(wb, config)
    build_tag_engine_sheet(wb, tags)
    build_qa_sheet(wb, rules)
    build_output_sheet(wb, output_columns)
    build_audit_log_sheet(wb)

    wb.save(output_path)

    pdf_name = config.get("source_pdf", "")
    print(f"\nGenerated : {output_path.resolve()}")
    print(f"Sheets    : {', '.join(wb.sheetnames)}")
    print(f"Tags      : {len(tags)}")
    print(f"PDF in tag: {pdf_name}")
    print()
    print("Next steps:")
    print(f"  1. Open {output_path.name} in Excel (DataSnipper auto-processes on open)")
    print(f"  2. Confirm {pdf_name} is in the same folder as the workbook")
    print("  3. Check TAG_ENGINE > SourceTag column - DataSnipper reads these cells")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a DataSnipper_(ds) workbook with tags built from your PDF",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive: reads PDF and prompts you to select search terms
  python scripts/generate_workbook.py --source-pdf "C:\\Audits\\invoice.pdf"

  # Non-interactive: pass terms directly
  python scripts/generate_workbook.py --source-pdf "invoice.pdf" --terms "Invoice Number,Total Amount"

  # Extract from page 2 instead of page 1
  python scripts/generate_workbook.py --source-pdf "invoice.pdf" --page 2
""",
    )
    parser.add_argument("--source-pdf", required=True,
                        help="Path to the source PDF")
    parser.add_argument("--output", default=None,
                        help="Output xlsx path (defaults to engagement_(ds).xlsx next to the PDF)")
    parser.add_argument("--terms", default=None,
                        help="Comma-separated search terms to skip the interactive prompt")
    parser.add_argument("--pages", default="1",
                        help="Pages to search: single ('1'), range ('1-5'), or list ('1,3,5'). Default: 1")
    parser.add_argument("--engagement-id", default="ENG-2026-001")
    parser.add_argument("--client", default="Sample Client Corp")
    parser.add_argument("--auditor", default="Lead Auditor")
    args = parser.parse_args()

    # Parse --pages into a list of ints: "1-5" → [1,2,3,4,5], "1,3" → [1,3]
    def parse_pages(spec: str) -> list[int]:
        pages = []
        for part in spec.split(","):
            part = part.strip()
            if "-" in part:
                start, end = part.split("-", 1)
                pages.extend(range(int(start), int(end) + 1))
            else:
                pages.append(int(part))
        return pages

    try:
        pages = parse_pages(args.pages)
    except ValueError:
        print(f"ERROR: invalid --pages value '{args.pages}'. Use '1', '1-5', or '1,3,5'")
        sys.exit(1)

    source_pdf  = Path(args.source_pdf)
    pdf_filename = source_pdf.name

    # Resolve output path — default to same folder as PDF
    output_path = Path(args.output) if args.output else source_pdf.parent / "engagement_(ds).xlsx"
    if not output_path.stem.endswith("_(ds)"):
        output_path = output_path.parent / (output_path.stem + "_(ds).xlsx")
        print(f"Note: renamed output to {output_path.name} (DataSnipper requires _(ds) suffix)")

    # ── Determine search terms ────────────────────────────────────────────────
    tags: list[TagDefinition] | None = None

    if args.terms:
        # Non-interactive: terms passed directly on the command line
        term_list = [t.strip() for t in args.terms.split(",") if t.strip()]
        print(f"Using {len(term_list)} term(s) across {len(pages)} page(s) = {len(term_list) * len(pages)} tags")
        tags = make_dynamic_tags(pdf_filename, term_list, pages=pages)

    elif source_pdf.exists():
        # Interactive: read the PDF and let the user pick
        # Use first page in the range for candidate extraction
        print(f"Reading PDF: {source_pdf}")
        candidates = extract_pdf_candidates(source_pdf, page=pages[0])

        if candidates:
            selected = pick_terms_interactively(candidates)
            if selected:
                tags = make_dynamic_tags(pdf_filename, selected, pages=pages)
            else:
                print("No terms selected — using placeholder tags")
        else:
            print("No text extracted from PDF (may be a scanned image) — using placeholder tags")

    else:
        print(f"PDF not found at {source_pdf} — using placeholder tags")
        print("(This is expected if generating on a machine without the PDF)")

    if tags is None:
        tags = make_sample_ap_tags(source_doc=pdf_filename)

    # ── Build config and generate ─────────────────────────────────────────────
    today = date.today()
    config = {
        "engagement_id":   args.engagement_id,
        "engagement_type": "AP",
        "period_start":    today.replace(day=1, month=1),
        "period_end":      today,
        "lead_auditor":    args.auditor,
        "client_name":     args.client,
        "workflow_type":   "SIMPLE",
        "source_pdf":      pdf_filename,
    }

    generate(output_path, tags, config)


if __name__ == "__main__":
    main()
